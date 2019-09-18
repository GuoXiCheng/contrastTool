import time
import tkinter

import xlrd, xlwt,openpyxl,datetime,re
from openpyxl.styles import PatternFill, numbers
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from tkinter import filedialog
from tkinter.filedialog import askdirectory
from tkinter import *
from tkinter import messagebox
from xlutils.styles import Styles
import math,os

SOURCE_READROAD = "C:/Users/Administrator/Desktop/内部進捗（源文件）.xlsx";
TARGET_READROAD = "C:/Users/Administrator/Desktop/内部進捗（目标文件）.xlsx";
OUTPUT_ROAD = "C:/Users/Administrator/Desktop";
SHEET_NAME = "進捗明細";
START_ROW = 3;#从第几行开始遍历数据，起始位为0
TRUE_COLOR = "FF92D14F";
FALSE_COLOR = "FFFF0000";

sourceBook = openpyxl.load_workbook(SOURCE_READROAD);
sourceSheet = sourceBook[SHEET_NAME];

targetBook = openpyxl.load_workbook(TARGET_READROAD);
targetSheet = targetBook[SHEET_NAME];

outBook = Workbook();
outSheet = outBook["Sheet"];
outSheet.title = "進捗明細";
# outBook.save(OUTPUT_ROAD + "\\" + "内部進捗（对比报告）.xlsx");

#遍历第三行，将单元格地址与背景色按键值对存储
def cellColor():
    colList = [];
    for col in range(len(tuple(sourceSheet.columns))):
        cellIndex = get_column_letter(col + 1) + "3";
        colList.append(cellIndex);
    colDict = dict.fromkeys(colList);
    for index in range(len(colList)):
        colorIndex = sourceSheet[colList[index]].fill.start_color.index;
        if type(colorIndex) == type(1):
            colDict[colList[index]] = openpyxl.styles.colors.COLOR_INDEX[colorIndex];
        elif type(colorIndex) == type(""):
            colDict[colList[index]] = colorIndex;
    return colDict;

#根据单元格颜色的字典进行分类，每一组颜色相同的单元格为一组存储单元格地址
def colNum():
    # colList = list(cellColor().items());
    colDict = cellColor();
    colNumList = [];
    temp = [];
    for key in colDict:
        if len(temp) == 0 or colDict[key] == "00000000" or colDict[key] == colDict[temp[0]]:
            temp.append(key);
        else:
            colNumList.append(tuple(temp));
            temp.clear();
            temp.append(key);
    return colNumList;

#对单元格地址进行切片，取得相应的列号
def getColList():
    colNumList = colNum();
    colList = [];
    temp = [];
    for i in range(len(colNumList)):
        for j in range(len(colNumList[i])):
            temp.append(((colNumList[i])[j])[0:-1]);
        colList.append(tuple(temp));
        temp.clear();
    return colList;


#对单元格地址进行遍历，判断单元格中的字段，取得单元格中为“実績”、“作業時間（H）”、“進捗率”的列号
def getContrastCol():
    colNumList = getColList();
    contrastColList = [];
    temp = [];
    for colTup in colNumList:
        for col in colTup:
            if sourceSheet[col + "3"].value == "実績" or sourceSheet[col + "2"].value == "作業時間（H）" or sourceSheet[col + "2"].value == "進捗率":
                temp.append(col);
        contrastColList.append(tuple(temp));
        temp.clear();
    return contrastColList;

#接受一个包含列号的元祖参数，将该列的所有数据读入列表中
def getData(colTup,sheet):
    dataList = [colTup];
    temp = [];
    for index in range(START_ROW,len(tuple(sheet.rows))):#从第四行开始遍历
        for col in colTup:
            temp.append(sheet[col + str(index + 1)].value);
        dataList.append(tuple(temp));
        temp.clear();
    return dataList;

#调试方式：换行打印列表
def printList(list):
    for param in list:
        print(param);

#传入两个数据列表，一个是源文件的，一个是目标文件的，进行比较后生成第三个列表
def contrastData(oList,cList):
    resultList = [oList[0]];
    temp = [];
    for i in range(1,len(oList)):
        for j in range(len(oList[i])):
            if (oList[i])[j] == (cList[i])[j]:
                temp.append("true");
            else:
                temp.append("false");
        resultList.append(tuple(temp));
        temp.clear();
    return resultList

#根据结果列表去给目标文件标记颜色，并返回数据不同的行号
def markTarget(rList):
    trueFill = PatternFill(fill_type='solid', fgColor=TRUE_COLOR);
    falseFill = PatternFill(fill_type='solid', fgColor=FALSE_COLOR);
    colList = rList[0];
    rList = rList[1:];
    differRow = [];
    for i in range(len(rList)):
        for j in range(len(rList[i])):
            if (rList[i])[j] == "true":
                targetSheet[colList[j] + str(START_ROW + i + 1)].fill = trueFill;
            elif (rList[i])[j] == "false":
                targetSheet[colList[j] + str(START_ROW + i + 1)].fill = falseFill;
                differRow.append(START_ROW + i + 1);
    targetBook.save(TARGET_READROAD);
    differRow = list((set(differRow)));
    differRow.sort();
    return differRow;

#传入一个列的元祖，返回该列所属的标题
def getTitle(colTup):
    colList = getColList()[1:];
    titleList = [];
    temp = [];
    for cols in colList:
        for col in cols:
            cellValue = sourceSheet[col + "2"].value;
            if cellValue != None and not "担当" in cellValue:
                cellValue = ''.join(cellValue.split());
                temp.append(cellValue);
        titleList.append(tuple(temp));
        temp.clear();
    title = "";
    for index in range(len(colList)):
        if colList[index].__contains__(colTup[0]):
            return titleList[index];
    return None;

#取得数据不同的行号，抽取数据写入一个新的excel
def writeToXL(rowList,col,title,startRow):#抽出数据的行号；抽出数据的列号；抽出数据的标题；从第几行开始写入
    # title = getTitle()[0];
    # col = getContrastCol()[1];
    # print(col);
    # print(rowList);
    if len(rowList) != 0:
        for i in range(len(rowList)):#有几组不同数据就要循环几次
            for j in range(len(title)):#首先循环写入标题
                titleIndex = get_column_letter(3 + j) + str(i + i + 1 + startRow);
                outSheet[titleIndex] = title[j];
            handleNameIndex = "A" + str(i + i + 2 + startRow);
            titleNameIndex = "B" + str(i + i + 2 + startRow);
            outSheet[handleNameIndex] = targetSheet["D" + str(rowList[i])].value;
            outSheet[titleNameIndex] = getObjectName(col[0]);
            for k in range(len(col)):#再循环写入数据
                dataIndex = get_column_letter(3 + k) + str( i + i + 2 + startRow);
                cellValue = targetSheet[col[k] + str(rowList[i])].value;
                if isinstance(cellValue,float) :
                    outSheet[dataIndex].number_format = numbers.FORMAT_PERCENTAGE;
                if isinstance(cellValue, datetime.datetime):
                    # outSheet[dataIndex].number_format = numbers.FORMAT_DATE_DATETIME;
                    time= str(cellValue).split("-");
                    day = time[2].split(" ");
                    outSheet[dataIndex] = time[0] +"/" + time[1] + "/" + day[0];
                else:
                    outSheet[dataIndex] = cellValue;




        # for i in range(len(title)):
        #     titleIndex = get_column_letter(startRow + 2 + i) + "1";
        #     outSheet[titleIndex] = title[i];
        # for j in range(len(rowList)):
        #     handleNameIndex = get_column_letter(startRow) + str(j + 2);
        #     titleNameIndex = get_column_letter(startRow + 1) + str(j + 2);
        #     outSheet[handleNameIndex] = targetSheet["D" + str(rowList[j])].value;
        #     outSheet[titleNameIndex] = getObjectName(col[0]);
        #     for k in range(len(col)):
        #         cellIndex = get_column_letter(startRow + 2 + k) + str(j + 2);
        #         outSheet[cellIndex] = targetSheet[col[k] + str(rowList[j])].value;
    outBook.save(OUTPUT_ROAD + "/" + "内部進捗（对比报告）.xlsx");

#传入列号，遍历取得大标题的名字
def getObjectName(colEn):
    colNum = column_index_from_string(colEn);
    while targetSheet[colEn + "1"].value == None:
        colNum = colNum - 1;
        colEn = get_column_letter(colNum);
    objectName =  "".join((targetSheet[colEn + "1"].value).split());
    return objectName.split("（")[0];

# contrastCol = getContrastCol();
# oList = getData(dataCol[1],sourceSheet);
# cList = getData(dataCol[1],targetSheet);
# differRow = markTarget(contrastData(oList,cList));
# writeToXL(differRow,0);

differNum = 0;

#主程序
def main():
    contrastCol = getContrastCol();
    listLength = 0;
    global differNum;
    for colTup in contrastCol:
        oList = getData(colTup,sourceSheet);
        cList = getData(colTup,targetSheet);
        differRow = markTarget(contrastData(oList,cList));
        if len(differRow) != 0:
            writeToXL(differRow, colTup, getTitle(colTup), listLength);
            listLength = len(differRow) * 2 + listLength;
        differNum = differNum + len(differRow);
    en = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"];
    for e in en:
        outSheet.column_dimensions[e].width = 25;
    outBook.save(OUTPUT_ROAD + "/" + "内部進捗（对比报告）.xlsx");


#GUI界面
def startRun():
    SOURCE_READROAD = sourceE.get();
    TARGET_READROAD = targetE.get();
    OUTPUT_ROAD = outputE.get();
    if os.path.exists(SOURCE_READROAD) and os.path.exists(TARGET_READROAD) and os.path.exists(OUTPUT_ROAD):
        main();
        messagebox.showinfo("message","运行完成\n" + "共抽出" + str(differNum) + "条数据");
        os._exit(0);
    else:
        messagebox.showwarning("Warning","路径不存在!");
    return ;


def selectSource():
    sourceRoad = filedialog.askopenfilename();
    sourceE.set(sourceRoad);

def selectTarget():
    targetRoad = filedialog.askopenfilename();
    targetE.set(targetRoad);

def selectOutput():
    outputRoad = askdirectory();
    outputE.set(outputRoad);

root = Tk();
root.title("对比工具");
root.geometry("550x150+400+300")

sourceE = StringVar();
sourceFile = Label(root, text = "原始文件路径").grid(row = 0, column = 1);
sourceEntry = Entry(root, width = 55, textvariable = sourceE ).grid(row = 0, column = 3);
sourceButton = Button(root, text = "选择",width = 10, height = 1, command = selectSource).grid(row = 0, column = 5);

targetE = StringVar();
targetFile = Label(root, text = "目标文件路径").grid(row = 1, column = 1);
targetEntry = Entry(root, width = 55, textvariable = targetE).grid(row = 1, column = 3);
targetButton = Button(root, text = "选择",width = 10, height = 1, command = selectTarget).grid(row = 1, column = 5);

outputE = StringVar();
outputFile = Label(root, text = "输出文件路径").grid(row = 2, column = 1);
outputEntry = Entry(root, width = 55, textvariable = outputE).grid(row = 2, column = 3);
outputButton = Button(root, text = "选择",width = 10, height = 1, command = selectOutput).grid(row = 2, column = 5);

startButton = Button(root, text = "开始",width = 10, height = 1, command = startRun).grid(row = 3, column = 3);

root.mainloop();











# originalBook = xlrd.open_workbook(ORIGINAL_READROAD);
# originalSheet = originalBook.sheets()[0];#第一个sheet
# o_rows = originalSheet.nrows;
# o_cols = originalSheet.ncols;

# book = xlrd.open_workbook(ORIGINAL_READROAD,formatting_info=1);
# sheet = book.sheets()[0];
# s = Styles(book);
# print(s[sheet.cell(0,0)])


# for row in range(o_cols):
#     print(originalSheet.col_values(row));
#
# contrastBook = xlrd.open_workbook(CONTRAST_READROAD);
# contrastSheet = contrastBook.sheets()[0];
# for row in range(o_rows):
#     print("第" + str(row + 1) + "行", end = "");
#     for col in range(o_cols):
#         if originalSheet.cell_type(row,col) == 0:
#             print("\t空的",end = "");
#         else:
#             print("\t", end = "");
#             print(originalSheet.cell_value(row,col),end = "");
#             print("\t",end = "");
#     print()

# list = [];
# for row in range(o_rows):
#     print("第" + str(row + 1) + "行", end = "");
#     for col in range(o_cols):
#         if originalSheet.cell_type(row,col) == 0:
#             print("\t空的",end = "");
#         else:
#             print("\t", end = "");
#             print(originalSheet.cell_value(row,col) == contrastSheet.cell_value(row,col),end = "");
#             if(originalSheet.cell_value(row ,col ) != contrastSheet.cell_value(row,col)):
#                 temp = (row + 1,col + 1);
#                 list.append(temp);
#             # print("\t",end = "");
#     print()

# loadWorkBook = openpyxl.load_workbook(ORIGINAL_READROAD);
# loadSheet = loadWorkBook["進捗明細"];
# fill = PatternFill(fill_type='solid',fgColor='FFFF0000');


# print(list);
# for i in range(len(list)):
#     # print(EnList[((list[i])[1] - 1)] + str((list[i])[0]),end = "");
#     cell = loadSheet[EnList[((list[i])[1] - 1)] + str((list[i])[0])];
#     cell.fill = fill;
#
#
# loadWorkBook.save(CONTRAST_READROAD);

# print(openpyxl.styles.colors.COLOR_INDEX[loadSheet["AC3"].fill.start_color.index])
# print(loadSheet["P3"].fill.start_color.index)



